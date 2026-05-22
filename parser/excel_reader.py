import csv
import os


class ExcelReader:
    HEADER_INDICATORS = [
        "varname", "variable", "varlabel", "label", "vartype", "type",
        "varlen", "length", "origin", "codelist", "algorithm"
    ]

    def read(self, filepath: str) -> dict:
        if filepath.endswith(".csv"):
            return self._read_csv(filepath)
        else:
            return self._read_xlsx(filepath)

    def _read_csv(self, filepath: str) -> dict:
        filename = os.path.splitext(os.path.basename(filepath))[0]
        rows = []
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                stripped = {k.strip(): v.strip() for k, v in row.items()}
                rows.append(stripped)
        domain = self._extract_domain_from_filename(filename)
        return {domain: rows}

    def _read_xlsx(self, filepath: str) -> dict:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True)
        sheets = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []

            header_row_idx = self._find_header_row(ws)
            headers = []

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == header_row_idx:
                    headers = [str(c).strip() if c else "" for c in row]
                elif i > header_row_idx:
                    if any(c is not None for c in row):
                        row_dict = {}
                        for j, header in enumerate(headers):
                            if header:
                                val = row[j] if j < len(row) else ""
                                row_dict[header] = self._convert_value(val)
                        if row_dict:
                            rows.append(row_dict)

            sheets[sheet_name] = rows
        wb.close()
        return sheets

    def _find_header_row(self, ws) -> int:
        """Find the row index containing SPEC column headers."""
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            row_lower = [str(c).lower().strip() if c else "" for c in row]
            match_count = sum(1 for indicator in self.HEADER_INDICATORS
                              if any(indicator in cell for cell in row_lower))
            if match_count >= 3:
                return i
        return 0

    def _convert_value(self, val):
        """Convert cell value to appropriate string representation."""
        if val is None:
            return ""
        if isinstance(val, (int, float)):
            if isinstance(val, float) and val == int(val):
                return str(int(val))
            return str(val)
        return str(val).strip()

    def _extract_domain_from_filename(self, filename: str) -> str:
        parts = filename.replace("_spec", "").split("_")
        for part in parts:
            if part.upper() in {"AE", "DM", "CM", "LB", "VS", "EX", "MH", "EG", "PE", "SUPPAE", "SUPPDM"}:
                return part.upper()
        return filename.upper()

    def detect_domain(self, sheets: dict) -> str:
        return list(sheets.keys())[0]
