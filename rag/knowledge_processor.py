"""
Knowledge Processor - 知识库预处理模块

解析知识库原始文件，生成结构化知识块供向量化使用。
"""

import os
import re
import json
from dataclasses import dataclass, field, asdict
from typing import Optional
from pathlib import Path


@dataclass
class KnowledgeChunk:
    """知识块数据结构"""
    id: str
    type: str  # sas_code | macro | spec_variable | sdtm_ig_section | sdtm_rule | sdtm_domain | sdtm_concept
    source: str
    content: str
    domain: Optional[str] = None
    metadata: dict = field(default_factory=dict)
    """知识块数据结构"""
    id: str
    type: str  # sas_code | macro | spec_variable
    source: str
    content: str
    domain: Optional[str] = None
    metadata: dict = field(default_factory=dict)


class SASCodeParser:
    """SAS 代码解析器 - 提取有意义的代码块"""

    def parse_file(self, filepath: str) -> list[KnowledgeChunk]:
        """解析 SAS 文件，提取代码块"""
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read()

        filename = os.path.basename(filepath)
        domain = self._extract_domain(filename)

        chunks = []

        # 提取宏定义
        macro_chunks = self._extract_macros(content, filename, domain)
        chunks.extend(macro_chunks)

        # 提取 DATA 步
        data_chunks = self._extract_data_steps(content, filename, domain)
        chunks.extend(data_chunks)

        # 提取 PROC 步
        proc_chunks = self._extract_proc_steps(content, filename, domain)
        chunks.extend(proc_chunks)

        # 提取衍生变量逻辑
        derivation_chunks = self._extract_derivation_logic(content, filename, domain)
        chunks.extend(derivation_chunks)

        return chunks

    def _extract_domain(self, filename: str) -> str:
        """从文件名提取域"""
        name = filename.lower().replace('.sas', '')
        domain_map = {
            'ae': 'AE', 'dm': 'DM', 'cm': 'CM', 'lb': 'LB',
            'vs': 'VS', 'ex': 'EX', 'mh': 'MH', 'eg': 'EG',
            'pe': 'PE', 'ds': 'DS', 'sv': 'SV', 'ie': 'IE',
            'qs': 'QS', 'rs': 'RS', 'tr': 'TR', 'tu': 'TU',
            'pr': 'PR', 'fa': 'FA', 'co': 'CO', 'dd': 'DD',
            'dv': 'DV', 'se': 'SE', 'ss': 'SS', 'su': 'SU',
            'relrec': 'RELREC', 'trial': 'TRIAL', 'pc': 'PC',
            'is': 'IS', 'pf': 'PF', 'cv': 'CV', 'xo': 'XO',
            'mi': 'MI', 'ec': 'EC', 'ta': 'TA', 'te': 'TE',
            'td': 'TD', 'ti': 'TI', 'ts': 'TS', 'tv': 'TV',
        }
        for key, value in domain_map.items():
            if name == key or name.startswith(key + '_'):
                return value
        return name.upper()[:2]

    def _extract_macros(self, content: str, filename: str, domain: str) -> list[KnowledgeChunk]:
        """提取宏定义"""
        chunks = []
        pattern = r'%macro\s+(\w+)\s*\(([^)]*)\)\s*;(.*?)%mend\s*(?:\1)?\s*;'
        matches = re.finditer(pattern, content, re.IGNORECASE | re.DOTALL)

        for i, match in enumerate(matches):
            macro_name = match.group(1)
            params = match.group(2)
            body = match.group(3).strip()

            chunk = KnowledgeChunk(
                id=f"{domain.lower()}_macro_{macro_name}_{i}",
                type="macro",
                source=filename,
                domain=domain,
                content=f"%macro {macro_name}({params});\n{body}\n%mend;",
                metadata={
                    "macro_name": macro_name,
                    "parameters": self._parse_params(params),
                    "body_lines": len(body.split('\n'))
                }
            )
            chunks.append(chunk)

        return chunks

    def _extract_data_steps(self, content: str, filename: str, domain: str) -> list[KnowledgeChunk]:
        """提取 DATA 步"""
        chunks = []
        # 简化模式：匹配 data ... run;
        pattern = r'data\s+([^;]+);\s*(.*?)\s*run;'
        matches = re.finditer(pattern, content, re.IGNORECASE | re.DOTALL)

        for i, match in enumerate(matches):
            data_name = match.group(1).strip()
            body = match.group(2).strip()

            if len(body) < 50:  # 跳过太短的代码块
                continue

            chunk = KnowledgeChunk(
                id=f"{domain.lower()}_data_{i}",
                type="sas_code",
                source=filename,
                domain=domain,
                content=f"data {data_name};\n{body}\nrun;",
                metadata={
                    "step_type": "DATA",
                    "dataset_name": data_name,
                    "variables": self._extract_variables(body)
                }
            )
            chunks.append(chunk)

        return chunks

    def _extract_proc_steps(self, content: str, filename: str, domain: str) -> list[KnowledgeChunk]:
        """提取 PROC 步"""
        chunks = []
        pattern = r'proc\s+(\w+)\s+([^;]+);\s*(.*?)\s*(?:run|quit);'
        matches = re.finditer(pattern, content, re.IGNORECASE | re.DOTALL)

        for i, match in enumerate(matches):
            proc_name = match.group(1).upper()
            options = match.group(2).strip()
            body = match.group(3).strip()

            if len(body) < 30:
                continue

            chunk = KnowledgeChunk(
                id=f"{domain.lower()}_proc_{proc_name}_{i}",
                type="sas_code",
                source=filename,
                domain=domain,
                content=f"proc {proc_name} {options};\n{body}\nrun;",
                metadata={
                    "step_type": "PROC",
                    "proc_name": proc_name,
                    "options": options
                }
            )
            chunks.append(chunk)

        return chunks

    def _extract_derivation_logic(self, content: str, filename: str, domain: str) -> list[KnowledgeChunk]:
        """提取衍生变量逻辑（if-then-else, select-when 等）"""
        chunks = []

        # 提取复杂的 if-then-else 块
        pattern = r'(if\s+.*?\s+then\s+do;.*?end;)'
        matches = re.finditer(pattern, content, re.IGNORECASE | re.DOTALL)

        for i, match in enumerate(matches):
            logic = match.group(1).strip()

            variables = self._extract_variables(logic)

            chunk = KnowledgeChunk(
                id=f"{domain.lower()}_derivation_{i}",
                type="sas_code",
                source=filename,
                domain=domain,
                content=logic,
                metadata={
                    "logic_type": "conditional",
                    "variables": variables,
                    "origin": "Derived"
                }
            )
            chunks.append(chunk)

        return chunks

    def _parse_params(self, params_str: str) -> list[str]:
        """解析宏参数"""
        if not params_str.strip():
            return []
        params = []
        for p in params_str.split(','):
            p = p.strip()
            if p:
                # 提取参数名（去掉默认值）
                param_name = p.split('=')[0].strip()
                params.append(param_name)
        return params

    def _extract_variables(self, code: str) -> list[str]:
        """从代码中提取变量名"""
        # 匹配 SAS 变量名模式：字母开头，包含字母数字下划线
        pattern = r'\b([A-Z][A-Z0-9_]*)\b'
        matches = re.findall(pattern, code.upper())

        # 过滤 SAS 关键字
        sas_keywords = {
            'IF', 'THEN', 'ELSE', 'DO', 'END', 'AND', 'OR', 'NOT',
            'DATA', 'SET', 'MERGE', 'BY', 'WHERE', 'KEEP', 'DROP',
            'LENGTH', 'FORMAT', 'INFORMAT', 'LABEL', 'ATTRIB',
            'INPUT', 'PUT', 'RETURN', 'OUTPUT', 'RUN', 'QUIT',
            'PROC', 'SELECT', 'WHEN', 'OTHERWISE', 'END',
        }

        variables = [v for v in matches if v not in sas_keywords and len(v) >= 2]
        return list(set(variables))[:20]  # 限制数量


class MacroParser:
    """Macro 解析器"""

    def parse_file(self, filepath: str) -> list[KnowledgeChunk]:
        """解析 Macro 文件"""
        with open(filepath, 'r', encoding='utf-8', errors='replace') as f:
            content = f.read()

        filename = os.path.basename(filepath)
        macro_name = filename.replace('.txt', '').replace('.sas', '')

        # 提取宏文档（开头的注释）
        doc_match = re.match(r'/\*(.*?)\*/', content, re.DOTALL)
        doc_text = doc_match.group(1).strip() if doc_match else ""

        # 提取宏定义
        macro_match = re.search(r'%macro\s+(\w+)\s*\(([^)]*)\)', content, re.IGNORECASE)
        params = macro_match.group(2) if macro_match else ""

        chunk = KnowledgeChunk(
            id=f"macro_{macro_name}",
            type="macro",
            source=filename,
            content=content,
            metadata={
                "macro_name": macro_name,
                "description": self._extract_description(doc_text),
                "parameters": self._parse_params(params),
                "usage": self._extract_usage(doc_text)
            }
        )

        return [chunk]

    def _extract_description(self, doc: str) -> str:
        """从文档中提取描述"""
        lines = doc.split('\n')
        for line in lines:
            line = line.strip()
            if line.lower().startswith('description'):
                return line.split(':', 1)[-1].strip()
            if 'short description' in line.lower():
                return line.split(':', 1)[-1].strip()
        return ""

    def _extract_usage(self, doc: str) -> str:
        """从文档中提取用法示例"""
        lines = doc.split('\n')
        for i, line in enumerate(lines):
            if 'sample call' in line.lower() or 'usage' in line.lower():
                # 返回后续几行
                return '\n'.join(lines[i:i+3])
        return ""

    def _parse_params(self, params_str: str) -> list[str]:
        if not params_str.strip():
            return []
        params = []
        for p in params_str.split(','):
            p = p.strip()
            if p:
                param_name = p.split('=')[0].strip()
                params.append(param_name)
        return params


class SpecParser:
    """SPEC 模板解析器"""

    def parse_file(self, filepath: str, knowledge_base_parser=None) -> list[KnowledgeChunk]:
        """解析 SPEC Excel 文件"""
        import openpyxl

        wb = openpyxl.load_workbook(filepath, data_only=True)
        filename = os.path.basename(filepath)
        domain = filename.replace('.xlsx', '').upper()

        chunks = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 找到表头行
            header_row = self._find_header_row(ws)
            if header_row is None:
                continue

            headers = [cell.value for cell in ws[header_row]]
            header_map = {str(h).strip().lower(): i for i, h in enumerate(headers) if h}

            # 遍历数据行
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
                if not row[0]:  # 跳过空行
                    continue

                var_name = self._get_col_value(row, header_map, 'varname', 'variable')
                var_label = self._get_col_value(row, header_map, 'varlabel', 'label')
                origin = self._get_col_value(row, header_map, 'origin')
                algorithm = self._get_col_value(row, header_map, 'algorithm for programming', 'algorithm')

                if not var_name:
                    continue

                # 只处理有算法描述的衍生变量
                if algorithm and str(algorithm).strip():
                    chunk = KnowledgeChunk(
                        id=f"{domain.lower()}_spec_{var_name}_{row_idx}",
                        type="spec_variable",
                        source=filename,
                        domain=domain,
                        content=f"Variable: {var_name}\nLabel: {var_label}\nOrigin: {origin}\nAlgorithm: {algorithm}",
                        metadata={
                            "variable_name": var_name,
                            "variable_label": var_label,
                            "origin": origin,
                            "algorithm": str(algorithm).strip(),
                            "sheet": sheet_name,
                            "row": row_idx
                        }
                    )
                    chunks.append(chunk)

        wb.close()
        return chunks

    def _find_header_row(self, ws) -> Optional[int]:
        """找到包含变量名表头的行"""
        header_indicators = ['varname', 'variable', 'varlabel', 'label']

        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            row_lower = [str(c).lower().strip() if c else "" for c in row]
            match_count = sum(1 for ind in header_indicators
                              if any(ind in cell for cell in row_lower))
            if match_count >= 2:
                return i
        return None

    def _get_col_value(self, row: tuple, header_map: dict, *col_names) -> str:
        """获取列值"""
        for name in col_names:
            name_lower = name.lower()
            if name_lower in header_map:
                idx = header_map[name_lower]
                if idx < len(row) and row[idx]:
                    return str(row[idx])
        return ""


class KnowledgeProcessor:
    """知识库处理主类"""

    def __init__(self, knowledge_base_path: str = None):
        self.sas_parser = SASCodeParser()
        self.macro_parser = MacroParser()
        self.spec_parser = SpecParser()
        self.sdtm_ig_parser = None  # 延迟加载

        if knowledge_base_path:
            self.knowledge_base_path = Path(knowledge_base_path)
        else:
            self.knowledge_base_path = Path("D:/Claude code/Knowlegde base")

    def _get_sdtm_ig_parser(self):
        """获取 SDTM IG 解析器（延迟加载）"""
        if self.sdtm_ig_parser is None:
            from rag.sdtm_ig_parser import SDTMIGParser
            self.sdtm_ig_parser = SDTMIGParser()
        return self.sdtm_ig_parser

    def process_all(self, output_dir: str = None) -> dict:
        """处理所有知识库文件"""
        all_chunks = []

        # 处理 SAS 代码
        sas_dir = self.knowledge_base_path / "SAS code"
        if sas_dir.exists():
            for sas_file in sas_dir.glob("*.sas"):
                chunks = self.sas_parser.parse_file(str(sas_file))
                all_chunks.extend(chunks)
                print(f"Processed {sas_file.name}: {len(chunks)} chunks")

        # 处理 Macros
        macro_dir = self.knowledge_base_path / "SAS macro"
        if macro_dir.exists():
            for macro_file in macro_dir.glob("*.*"):
                chunks = self.macro_parser.parse_file(str(macro_file))
                all_chunks.extend(chunks)
                print(f"Processed {macro_file.name}: {len(chunks)} chunks")

        # 处理 SPEC 模板
        spec_dir = self.knowledge_base_path / "SPEC template"
        if spec_dir.exists():
            for spec_file in spec_dir.glob("*.xlsx"):
                chunks = self.spec_parser.parse_file(str(spec_file))
                all_chunks.extend(chunks)
                print(f"Processed {spec_file.name}: {len(chunks)} chunks")

        # 处理 SDTM IG
        sdtm_ig_dir = self.knowledge_base_path / "SDTM IG"
        if sdtm_ig_dir.exists():
            parser = self._get_sdtm_ig_parser()
            for pdf_file in sdtm_ig_dir.glob("*.pdf"):
                try:
                    chunks = parser.parse_file(str(pdf_file))
                    all_chunks.extend(chunks)
                    print(f"Processed {pdf_file.name}: {len(chunks)} chunks")
                except Exception as e:
                    print(f"Error processing {pdf_file.name}: {e}")

        # 保存结果
        if output_dir:
            output_path = Path(output_dir)
            output_path.mkdir(parents=True, exist_ok=True)

            # 按类型分组保存
            by_type = {}
            for chunk in all_chunks:
                if chunk.type not in by_type:
                    by_type[chunk.type] = []
                by_type[chunk.type].append(asdict(chunk))

            for chunk_type, items in by_type.items():
                output_file = output_path / f"{chunk_type}_chunks.json"
                with open(output_file, 'w', encoding='utf-8') as f:
                    json.dump(items, f, ensure_ascii=False, indent=2)
                print(f"Saved {len(items)} {chunk_type} chunks to {output_file}")

        return {
            "total_chunks": len(all_chunks),
            "by_type": {t: len(c) for t, c in by_type.items()} if output_dir else {}
        }

    def get_chunks(self) -> list[KnowledgeChunk]:
        """获取所有知识块（不保存到文件）"""
        all_chunks = []

        sas_dir = self.knowledge_base_path / "SAS code"
        if sas_dir.exists():
            for sas_file in sas_dir.glob("*.sas"):
                chunks = self.sas_parser.parse_file(str(sas_file))
                all_chunks.extend(chunks)

        macro_dir = self.knowledge_base_path / "SAS macro"
        if macro_dir.exists():
            for macro_file in macro_dir.glob("*.*"):
                chunks = self.macro_parser.parse_file(str(macro_file))
                all_chunks.extend(chunks)

        # 添加 SDTM IG
        sdtm_ig_dir = self.knowledge_base_path / "SDTM IG"
        if sdtm_ig_dir.exists():
            parser = self._get_sdtm_ig_parser()
            for pdf_file in sdtm_ig_dir.glob("*.pdf"):
                try:
                    chunks = parser.parse_file(str(pdf_file))
                    all_chunks.extend(chunks)
                except Exception as e:
                    print(f"Error processing {pdf_file.name}: {e}")

        return all_chunks


if __name__ == "__main__":
    processor = KnowledgeProcessor()
    stats = processor.process_all(output_dir="d:/Claude code/sdtm_gen/knowledge")
    print(f"\nTotal chunks: {stats['total_chunks']}")
